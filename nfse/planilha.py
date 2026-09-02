# -*- coding: utf-8 -*-
"""Exporta a conferencia para uma planilha, para o contador conferir fora.

O `RELATORIO.txt` serve para olhar rapido, mas nao da para filtrar, somar
por seccao nem cruzar com o razao. Quem confere numero trabalha em planilha.

**Os totais sao formula, nunca valor calculado aqui.** Se o contador filtrar
ou apagar uma linha, a soma acompanha -- um numero fixo mentiria em silencio
justamente na conferencia. Pelo mesmo motivo os valores vao como numero, com
formato de moeda, e nao como texto "R$ 1.234,56": texto nao soma.
"""

from __future__ import annotations

import io
from datetime import datetime

# Uma unica fonte, do sistema, em tudo -- planilha de conferencia nao e lugar
# de personalidade tipografica.
FONTE = "Arial"
MOEDA = 'R$ #,##0.00'
CABECALHO_FUNDO = "1F3A4A"


def _estilo(ws, linha_cabecalho: int, colunas: list) -> None:
    """Cabecalho fixo, com largura por coluna e painel congelado."""
    from openpyxl.styles import Alignment, Font, PatternFill

    for indice, (titulo, largura) in enumerate(colunas, start=1):
        celula = ws.cell(row=linha_cabecalho, column=indice, value=titulo)
        celula.font = Font(name=FONTE, bold=True, color="FFFFFF", size=10)
        celula.fill = PatternFill("solid", fgColor=CABECALHO_FUNDO)
        celula.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[celula.column_letter].width = largura
    ws.row_dimensions[linha_cabecalho].height = 26
    # Congelar deixa o cabecalho visivel ao rolar 276 linhas.
    ws.freeze_panes = ws.cell(row=linha_cabecalho + 1, column=1)


def _corpo(ws, primeira: int, ultima: int) -> None:
    from openpyxl.styles import Font

    for linha in ws.iter_rows(min_row=primeira, max_row=ultima):
        for celula in linha:
            celula.font = Font(name=FONTE, size=10)


def gerar(resultado, unidade: dict, competencia: str,
          aliquota_iss: float = 0.0) -> bytes:
    """Monta a planilha da conferencia e devolve os bytes do .xlsx."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    livro = Workbook()

    # ---------------------------------------------------------------- resumo
    resumo = livro.active
    resumo.title = "Resumo"
    resumo["A1"] = "Conferência de NFS-e"
    resumo["A1"].font = Font(name=FONTE, bold=True, size=14)
    resumo["A2"] = unidade.get("razao_social", "")
    resumo["A3"] = "CNPJ %s · competência %s" % (
        unidade.get("cnpj", ""), competencia)
    resumo["A4"] = "Gerado em %s" % datetime.now().strftime("%d/%m/%Y %H:%M")
    for linha in ("A2", "A3", "A4"):
        resumo[linha].font = Font(name=FONTE, size=10, color="555555")
    resumo.column_dimensions["A"].width = 42
    resumo.column_dimensions["B"].width = 14
    resumo.column_dimensions["C"].width = 18

    linhas_resumo = [
        ("", "", ""),
        ("O que foi lido", "Quantidade", "Valor"),
        # As contagens apontam para as outras abas: mexer la reflete aqui.
        ("Lançamentos nos relatórios", resultado.total_lancamentos, None),
        ("Vão virar nota", "=COUNTA(Notas!A2:A100000)",
         "=SUM(Notas!F2:F100000)"),
        ("Travados, esperando correção", "=COUNTA(Travados!A2:A100000)",
         "=SUM(Travados!E2:E100000)"),
        ("Não geram nota (regra da clínica)",
         "=COUNTA('Não geram nota'!A2:A100000)",
         "=SUM('Não geram nota'!D2:D100000)"),
    ]
    for deslocamento, (rotulo, qtde, valor) in enumerate(linhas_resumo):
        linha = 6 + deslocamento
        resumo.cell(row=linha, column=1, value=rotulo)
        resumo.cell(row=linha, column=2, value=qtde)
        resumo.cell(row=linha, column=3, value=valor)
        for coluna in (1, 2, 3):
            celula = resumo.cell(row=linha, column=coluna)
            celula.font = Font(name=FONTE, size=10,
                               bold=(deslocamento == 1))
        if valor is not None:
            resumo.cell(row=linha, column=3).number_format = MOEDA

    aviso = resumo.cell(
        row=6 + len(linhas_resumo) + 1, column=1,
        value="Os totais são fórmulas: ao filtrar ou remover linhas nas "
              "outras abas, as somas acompanham.")
    aviso.font = Font(name=FONTE, size=9, italic=True, color="777777")
    aviso.alignment = Alignment(wrap_text=True)

    # ----------------------------------------------------------------- notas
    ws = livro.create_sheet("Notas")
    _estilo(ws, 1, [
        ("Data", 12), ("Paciente", 38), ("CPF/CNPJ", 18), ("Cidade", 20),
        ("Seção do caixa", 26), ("Valor", 14), ("ISS", 12),
        ("Caixa", 14), ("Lançamento", 16),
    ])
    for indice, nota in enumerate(resultado.notas, start=2):
        tomador = nota.tomador or {}
        ws.cell(row=indice, column=1, value=nota.data)
        ws.cell(row=indice, column=2, value=tomador.get("nome", ""))
        ws.cell(row=indice, column=3, value=tomador.get("documento", ""))
        ws.cell(row=indice, column=4, value="%s/%s" % (
            tomador.get("cidade", ""), tomador.get("uf", "")))
        ws.cell(row=indice, column=5, value=nota.secao)
        celula_valor = ws.cell(row=indice, column=6, value=float(nota.valor))
        celula_valor.number_format = MOEDA
        # O ISS sai de formula sobre o valor da propria linha: some com o
        # filtro e confere na hora, sem depender do que este script calculou.
        iss = ws.cell(row=indice, column=7, value="=F%d*$K$1" % indice)
        iss.number_format = MOEDA
        ws.cell(row=indice, column=8, value=nota.caixa)
        ws.cell(row=indice, column=9, value=nota.lancto)
    _corpo(ws, 2, len(resultado.notas) + 1)

    # A aliquota fica numa celula propria, nomeada -- assumption visivel, nao
    # numero enterrado dentro de cada formula.
    ws["K1"] = float(aliquota_iss) / 100.0
    ws["J1"] = "Alíquota ISS"
    ws["J1"].font = Font(name=FONTE, bold=True, size=9)
    ws["K1"].number_format = "0.00%"
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 10

    # ------------------------------------------------------------- travados
    ws = livro.create_sheet("Travados")
    _estilo(ws, 1, [
        ("Data", 12), ("Paciente no caixa", 36), ("Motivo", 34),
        ("O que fazer", 46), ("Valor", 14), ("Seção", 24), ("Lançamento", 16),
    ])
    for indice, p in enumerate(resultado.pendencias, start=2):
        ws.cell(row=indice, column=1, value=p.data)
        ws.cell(row=indice, column=2, value=p.nome)
        ws.cell(row=indice, column=3, value=p.titulo)
        ws.cell(row=indice, column=4, value=p.orientacao)
        celula = ws.cell(row=indice, column=5, value=float(p.valor))
        celula.number_format = MOEDA
        ws.cell(row=indice, column=6, value=p.secao)
        ws.cell(row=indice, column=7, value=p.lancto)
    _corpo(ws, 2, len(resultado.pendencias) + 1)

    # --------------------------------------------------------- nao emitidas
    ws = livro.create_sheet("Não geram nota")
    _estilo(ws, 1, [
        ("Motivo", 44), ("Seção do caixa", 30), ("Lançamentos", 14),
        ("Valor", 16),
    ])
    linha = 2
    for motivo, dados in sorted(resultado.descartes.items(),
                                key=lambda x: -x[1]["valor"]):
        for secao, sd in sorted(dados.get("secoes", {}).items(),
                                key=lambda x: -x[1]["valor"]):
            ws.cell(row=linha, column=1, value=motivo)
            ws.cell(row=linha, column=2, value=secao)
            ws.cell(row=linha, column=3, value=sd["qtde"])
            celula = ws.cell(row=linha, column=4, value=float(sd["valor"]))
            celula.number_format = MOEDA
            linha += 1
    _corpo(ws, 2, linha - 1)

    memoria = io.BytesIO()
    livro.save(memoria)
    return memoria.getvalue()
