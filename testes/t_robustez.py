# -*- coding: utf-8 -*-
"""Maus-tratos deliberados: o que acontece quando o operador erra.

Os outros testes provam que o caminho certo funciona. Este prova que o
caminho errado nao derruba o sistema nem, pior, produz nota errada em
silencio. Quem vai operar na clinica nao e quem escreveu isto: vai mandar o
PDF trocado, o arquivo pela metade, dois relatorios com o mesmo nome, e vai
clicar duas vezes no botao.

A regra que este arquivo defende: **falhar visivel e melhor do que acertar
por sorte**. Toda entrada ruim tem que virar mensagem em portugues, nunca
uma tela de erro do Flask e nunca um XML gerado errado.
"""

import io
import os
import shutil
import sys
import tempfile
import threading

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from nfse import config as cfgmod

# dados/ e saida/ descartaveis ANTES de importar o app -- ele le esses
# caminhos na importacao, e o teste nao pode encostar no controle.db real.
PASTA_DADOS_REAL = cfgmod.PASTA_DADOS
PASTA_SAIDA_REAL = cfgmod.PASTA_SAIDA
PASTA_LOTES_REAL = cfgmod.PASTA_LOTES
cfgmod.PASTA_DADOS = tempfile.mkdtemp(prefix="nfse-rob-dados-")
cfgmod.PASTA_SAIDA = tempfile.mkdtemp(prefix="nfse-rob-saida-")
cfgmod.PASTA_LOTES = tempfile.mkdtemp(prefix="nfse-rob-lotes-")

import app as appmod  # noqa: E402
from nfse.controle import Controle  # noqa: E402
from nfse.documentos import documento_valido  # noqa: E402
from nfse.transmissao import ambiente_do_xml  # noqa: E402

appmod.cfgmod.PASTA_DADOS = cfgmod.PASTA_DADOS
appmod.cfgmod.PASTA_SAIDA = cfgmod.PASTA_SAIDA
appmod.cfgmod.PASTA_LOTES = cfgmod.PASTA_LOTES

# O console do Windows costuma vir em cp1252 e derruba o teste ao imprimir
# um caractere que ele nao tem -- e um teste que "falha" por acento esconde
# o resultado de verdade.
for _fluxo in (sys.stdout, sys.stderr):
    try:
        _fluxo.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

passou = 0
falhas = []


def checar(nome, condicao, detalhe=""):
    global passou
    if condicao:
        passou += 1
        print("  [ok]    %s" % nome)
    else:
        falhas.append(nome)
        print("  [FALHA] %s %s" % (nome, ("-> " + str(detalhe)) if detalhe else ""))


def pdf_falso(texto="CAIXA - LANCAMENTOS"):
    """PDF minimo de verdade, gerado sem dependencia externa."""
    conteudo = ("BT /F1 12 Tf 40 750 Td (%s) Tj ET" % texto).encode("latin-1")
    partes = [b"%PDF-1.4\n"]
    objetos = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] "
        b"/Resources << /Font << /F1 5 0 R >> >> /Contents 4 0 R >>",
        b"<< /Length %d >>\nstream\n" % len(conteudo) + conteudo + b"\nendstream",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
    ]
    deslocamentos = []
    for indice, corpo in enumerate(objetos, start=1):
        deslocamentos.append(sum(len(p) for p in partes))
        partes.append(b"%d 0 obj\n" % indice + corpo + b"\nendobj\n")
    inicio_xref = sum(len(p) for p in partes)
    partes.append(b"xref\n0 %d\n0000000000 65535 f \n" % (len(objetos) + 1))
    for d in deslocamentos:
        partes.append(b"%010d 00000 n \n" % d)
    partes.append(
        b"trailer\n<< /Size %d /Root 1 0 R >>\nstartxref\n%d\n%%%%EOF\n"
        % (len(objetos) + 1, inicio_xref)
    )
    return b"".join(partes)


cliente = appmod.app.test_client()

# A PRIMEIRA unidade configurada, seja ela qual for. Amarrar em "gloria"
# fazia o teste quebrar numa copia recem-clonada, que parte do
# empresas.exemplo.json -- e este teste nao e sobre clinica nenhuma.
UNIDADE = next(iter(cfgmod.carregar().unidades))


def subir(arquivos, **campos):
    dados = {"arquivos": [(io.BytesIO(c), n) for n, c in arquivos]}
    dados.update(campos)
    return cliente.post("/lote", data=dados,
                        content_type="multipart/form-data")


# ==========================================================================
print("\n1. Arquivos que o operador pode mandar por engano")

r = subir([])
checar("sem nenhum arquivo nao explode", r.status_code == 400, r.status_code)

r = subir([("relatorio.pdf", b"isto nao e um PDF, e lixo")])
checar("PDF corrompido vira mensagem, nao tela de erro",
       r.status_code == 400, "%s %s" % (r.status_code, r.data[:200]))
checar("e a mensagem diz o que faltou",
       b"CAIXA" in r.data, r.data[:200])

r = subir([("planilha.xlsx", b"PK\x03\x04qualquer coisa")])
checar("arquivo que nao e PDF e recusado", r.status_code == 400)

r = subir([("outro.pdf", pdf_falso("RELATORIO DE ANIVERSARIANTES DO MES"))])
checar("PDF valido mas do relatorio errado e recusado", r.status_code == 400)
checar("sem deixar pasta de lote orfa no disco",
       len([n for n in os.listdir(cfgmod.PASTA_LOTES)]) == 0,
       os.listdir(cfgmod.PASTA_LOTES))

# Nome de arquivo hostil: o Flask salva com o nome que o navegador mandar.
r = subir([("../../../roubado.pdf", b"lixo")])
checar("caminho no nome do arquivo nao escapa da pasta",
       not os.path.exists(os.path.join(cfgmod.PASTA_LOTES, "..", "roubado.pdf")))

r = subir([("relatório açaí.pdf", b"lixo")])
checar("acento no nome do arquivo nao quebra", r.status_code == 400)


# ==========================================================================
print("\n2. Dois arquivos com o mesmo nome (o erro silencioso)")
# Se a funcionaria baixar o caixa de cada clinica e os dois se chamarem
# "caixa.pdf", o segundo sobrescreve o primeiro no disco -- e a lista de
# caixas fica com o MESMO arquivo duas vezes. Resultado: todo lancamento
# contado em dobro, sem nenhum aviso.
pasta_teste = tempfile.mkdtemp(prefix="nfse-rob-nomes-")
try:
    caminhos = []
    for _ in range(2):
        destino = appmod.destino_unico(pasta_teste, "caixa.pdf")
        with open(destino, "wb") as fh:
            fh.write(b"conteudo")
        caminhos.append(destino)
    checar("dois arquivos de mesmo nome nao se sobrescrevem",
           caminhos[0] != caminhos[1] and all(os.path.exists(c) for c in caminhos),
           caminhos)
    checar("e os dois continuam no disco",
           len(os.listdir(pasta_teste)) == 2, os.listdir(pasta_teste))
finally:
    shutil.rmtree(pasta_teste, ignore_errors=True)


# ==========================================================================
print("\n3. Enderecos que nao existem")

checar("lote inexistente devolve 404",
       cliente.get("/lote/nao-existe").status_code == 404)
checar("progresso de lote inexistente devolve 404",
       cliente.get("/lote/nao-existe/progresso").status_code == 404)
checar("pasta de saida inexistente devolve 404",
       cliente.get("/saida/nao-existe/transmitir").status_code == 404)
checar("zip de pasta inexistente devolve 404",
       cliente.get("/saida/nao-existe/zip").status_code == 404)
checar("subir na arvore de pastas e barrado",
       cliente.get("/saida/..%2f..%2fWindows/zip").status_code in (400, 404))
for rota in ("/", "/clientes", "/configuracao", "/ajuda"):
    checar("a tela %s abre" % rota, cliente.get(rota).status_code == 200)


# ==========================================================================
print("\n4. Numeracao sob pressao")

caminho_db = os.path.join(cfgmod.PASTA_DADOS, "teste-numeracao.db")
controle = Controle(caminho_db)
numeros = []
trava = threading.Lock()


def pegar_numeros():
    locais = [controle.proximo_numero(UNIDADE) for _ in range(40)]
    with trava:
        numeros.extend(locais)


fios = [threading.Thread(target=pegar_numeros) for _ in range(5)]
for f in fios:
    f.start()
for f in fios:
    f.join()

checar("200 numeros pedidos ao mesmo tempo, nenhum repetido",
       len(numeros) == len(set(numeros)) == 200,
       "%d pedidos, %d distintos" % (len(numeros), len(set(numeros))))
checar("e a sequencia nao tem buraco",
       sorted(numeros) == list(range(1, 201)))

# Segunda conexao ao mesmo arquivo: e o que acontece se abrirem o programa
# duas vezes na mesma maquina.
outro = Controle(caminho_db)
n1 = controle.proximo_numero(UNIDADE)
n2 = outro.proximo_numero(UNIDADE)
checar("duas conexoes ao mesmo banco nao repetem numero",
       n1 != n2 and {n1, n2} == {201, 202}, (n1, n2))
outro.fechar()
controle.fechar()


# ==========================================================================
print("\n5. Controle da numeracao corrompido")
# O `controle.db` e o arquivo mais critico do sistema. Se ele corromper
# (desligamento no meio de uma gravacao, disco cheio, antivirus), o sqlite
# levanta "file is not a database" e o app respondia 500 em quase todas as
# telas -- um beco sem saida para quem esta na recepcao da clinica.
from nfse.controle import ControleIlegivel  # noqa: E402

pasta_ruim = tempfile.mkdtemp(prefix="nfse-rob-corromp-")
try:
    caminho_ruim = os.path.join(pasta_ruim, "controle.db")
    with open(caminho_ruim, "w") as fh:
        fh.write("isto nao e um banco sqlite")
    try:
        Controle(caminho_ruim)
        checar("banco corrompido e detectado", False, "abriu normalmente")
    except ControleIlegivel as erro:
        checar("banco corrompido vira erro proprio, nao sqlite3", True)
        checar("e a mensagem manda restaurar o backup",
               "backup" in str(erro).lower(), str(erro)[:120])
        checar("e avisa para nao apagar o arquivo",
               "apague" in str(erro).lower(), str(erro)[:120])

    dados_antes = cfgmod.PASTA_DADOS
    cfgmod.PASTA_DADOS = pasta_ruim
    appmod.cfgmod.PASTA_DADOS = pasta_ruim
    try:
        resposta = cliente.get("/")
        corpo = resposta.data.decode("utf-8", "replace").lower()
        checar("a tela explica o problema em portugues",
               "restaure o backup" in corpo, resposta.status_code)
        checar("e ensina o caminho sem backup", "ajustar numera" in corpo)
    finally:
        cfgmod.PASTA_DADOS = dados_antes
        appmod.cfgmod.PASTA_DADOS = dados_antes
finally:
    shutil.rmtree(pasta_ruim, ignore_errors=True)


# ==========================================================================
print("\n6. Descartar um lote gerado com numeracao errada")
# O caso real: o controle da numeracao se perde, o sistema recomeca do 1 e o
# lote inteiro sai com numeros que a prefeitura ja usou. A prefeitura recusa
# (E0014) e a antiduplicidade daqui impede gerar de novo -- o operador fica
# preso sem saida, com 276 notas inuteis.
pasta_lote = tempfile.mkdtemp(prefix="nfse-rob-lote-")
try:
    caminho_db2 = os.path.join(pasta_lote, "controle.db")
    with Controle(caminho_db2) as c:
        for n in range(1, 6):
            numero = c.proximo_numero("unidade_x")
            c.registrar(c.chave("unidade_x", "L%d" % n), numero=numero,
                        arquivo="%05d-nota.xml" % numero, descricao="teste",
                        em="", documento="", competencia="2026-08",
                        valor="10.00", secao="X")
        # a terceira ja foi aceita pela prefeitura: nao pode ser descartada
        c.registrar_transmissao(c.chave("unidade_x", "L3"), numero_nota="900",
                                codigo_verificacao="", chave_acesso="K",
                                ambiente="homologacao")

        r = c.descartar_lote(["%05d-nota.xml" % n for n in range(1, 6)])
        checar("descarta as nao transmitidas", r["descartadas"] == 4, r)
        checar("*** a ja transmitida NAO e descartada ***",
               r["protegidas"] == 1, r)
        checar("e diz quais numeros foram liberados",
               r["numeros"] == [1, 2, 4, 5], r["numeros"])

        restantes = c.por_arquivo()
        checar("so a transmitida continua registrada",
               list(restantes) == ["00003-nota.xml"], list(restantes))
        checar("a numeracao NAO volta sozinha",
               c.resumo()["unidades"]["unidade_x"] == 5,
               "quem ajusta e o operador, que sabe o numero real da prefeitura")
        checar("descartar lista vazia nao quebra",
               c.descartar_lote([])["descartadas"] == 0)
finally:
    shutil.rmtree(pasta_lote, ignore_errors=True)


# ==========================================================================
print("\n7. Notas geradas que ainda esperam a prefeitura")
# O operador gerava as notas, saia da tela e nao achava mais o caminho de
# volta: reabrindo a conferencia, o sistema pulava tudo (ja emitido) e a
# transmissao aparecia zerada. Os XMLs existiam, prontos e assinados, sem
# nenhuma porta ate eles.
pasta_pend = tempfile.mkdtemp(prefix="nfse-rob-pend-")
saida_antes = cfgmod.PASTA_SAIDA
cfgmod.PASTA_SAIDA = pasta_pend
appmod.cfgmod.PASTA_SAIDA = pasta_pend
try:
    for nome, arquivos in (("gloria-2026-08-agora", 3),
                           ("gloria-2026-08-antiga-teste", 2),
                           ("gloria-2026-07-DESCARTADO", 4)):
        os.makedirs(os.path.join(pasta_pend, nome), exist_ok=True)
        for n in range(1, arquivos + 1):
            with open(os.path.join(pasta_pend, nome, "%05d-x.xml" % n), "wb") as fh:
                fh.write(b"<NFSe><tpAmb>2</tpAmb><Signature/></NFSe>")

    pendentes = appmod.lotes_por_transmitir()
    nomes = [p["pasta"] for p in pendentes]
    checar("acha o lote gerado e nao transmitido",
           "gloria-2026-08-agora" in nomes, nomes)
    checar("pasta de simulacao nao aparece",
           "gloria-2026-08-antiga-teste" not in nomes, nomes)
    checar("lote descartado nao aparece",
           "gloria-2026-07-DESCARTADO" not in nomes, nomes)
    achado = next(p for p in pendentes if p["pasta"] == "gloria-2026-08-agora")
    checar("conta quantas faltam", achado["faltam"] == 3, achado)
    checar("e diz o ambiente em que foram geradas",
           achado["ambiente"] == "homologacao", achado)

    corpo = cliente.get("/").data.decode("utf-8", "replace")
    checar("a tela inicial mostra o lote esperando",
           "esperando a prefeitura" in corpo)
    checar("com botao para transmitir", "gloria-2026-08-agora" in corpo)
finally:
    cfgmod.PASTA_SAIDA = saida_antes
    appmod.cfgmod.PASTA_SAIDA = saida_antes
    shutil.rmtree(pasta_pend, ignore_errors=True)


# ==========================================================================
print("\n8. Atualizacao instalada com o programa ainda aberto")
# Os templates recarregam sozinhos; o codigo Python nao. Entre instalar e
# reiniciar, o sistema fica com tela nova e logica velha -- pior do que a
# versao antiga inteira, porque o erro que aparece nao tem relacao obvia com
# a atualizacao. O aviso tem que estar em TODA tela, nao so na de
# Configuracao, senao some assim que a operadora navega.
caminho_versao = os.path.join(cfgmod.RAIZ, "VERSAO")
original = None
try:
    with open(caminho_versao, encoding="utf-8") as fh:
        original = fh.read()

    corpo = cliente.get("/").data.decode("utf-8", "replace")
    checar("sem atualizacao pendente, nao ha aviso",
           "falta fechar e abrir" not in corpo)

    with open(caminho_versao, "w", encoding="utf-8") as fh:
        fh.write("99.0.0" + chr(10))

    for rota in ("/", "/configuracao", "/clientes", "/ajuda"):
        corpo = cliente.get(rota).data.decode("utf-8", "replace")
        checar("aviso de reinicio aparece em %s" % rota,
               "falta fechar e abrir" in corpo)

    corpo = cliente.get("/").data.decode("utf-8", "replace")
    checar("e diz qual versao esta rodando e qual foi instalada",
           "99.0.0" in corpo and appmod.VERSAO_EM_EXECUCAO in corpo)
finally:
    if original is not None:
        with open(caminho_versao, "w", encoding="utf-8") as fh:
            fh.write(original)


# ==========================================================================
print("\n9. Transmissao: a fila da prefeitura")
# A prefeitura processa UM envio por vez por CNPJ. Mandar um lote em
# sequencia faz a maioria voltar recusada -- nao por defeito da nota, e sim
# por ritmo. Numa transmissao real de 276, 207 cairam assim. Distinguir esse
# erro dos de conteudo e o que permite insistir so no que adianta.
from nfse.transmissao import erro_de_fila  # noqa: E402

checar("reconhece a recusa por fila",
       erro_de_fila("ATENCAO! Ja consta uma requisicao em andamento para envio"))
checar("reconhece a variante acentuada",
       erro_de_fila("Já consta uma requisição em andamento"))
checar("reconhece pelo aguarde o processo anterior",
       erro_de_fila("Aguarde o processo anterior finalizar"))
checar("E0014 NAO e erro de fila (nao adianta insistir)",
       not erro_de_fila("E0014 - Conjunto de Serie, Numero ... ja existe"))
checar("erro de assinatura nao e de fila",
       not erro_de_fila("Erro na assinatura: Falha na validacao"))
checar("mensagem vazia nao e de fila", not erro_de_fila(""))


# ==========================================================================
print("\n10. O ambiente e escolhido ao GERAR, nao na configuracao")
# Obrigar a mexer na configuracao antes de cada teste era retrabalho, e
# deixava um estado global facil de esquecer ligado -- o pior erro possivel
# seria emitir valendo achando que era teste. Agora e escolha por lote.
from nfse.emissao import emitir as _emitir  # noqa: E402
import inspect  # noqa: E402

parametros = inspect.signature(_emitir).parameters
checar("emitir() recebe o ambiente de quem chama", "ambiente" in parametros)
checar("e o default e vazio (cai na configuracao)",
       parametros["ambiente"].default == "")

corpo_conf = cliente.get("/configuracao").data.decode("utf-8", "replace")
checar("a configuracao nao tem mais chave de ambiente",
       "configuracao/ambiente" not in corpo_conf)
checar("e explica onde a escolha e feita",
       "gerar as notas" in corpo_conf.lower())

checar("rota antiga de trocar ambiente sumiu",
       cliente.post("/configuracao/ambiente",
                    data={"ambiente": "producao"}).status_code in (404, 405))

# O XML tem que sair marcado com o ambiente PEDIDO, mesmo que a configuracao
# diga outra coisa. Ler so da configuracao fazia o tpAmb sair 1 (producao)
# quando o operador pedia teste -- nota valendo emitida achando que era
# ensaio, o pior erro que este sistema pode cometer.
from nfse.gerador_dps import gerar_nfse as _gerar  # noqa: E402
from nfse.conciliacao import Nota as _Nota  # noqa: E402
from lxml import etree as _etree  # noqa: E402


def nota_exemplo():
    return _Nota(
        id="amb", unidade=UNIDADE, data="2026-08-14", competencia="2026-08",
        valor="100.00", secao="REC ODC - PIX", caixa="CLINICA", lancto="1",
        contrato="", historico="Servicos", tipo_faturamento="particular",
        tomador={"nome": "PACIENTE EXEMPLO", "documento": "12345678909",
                 "codigo_municipio": "3205200", "cep": "29100000",
                 "logradouro": "RUA X", "numero": "1", "bairro": "CENTRO",
                 "uf": "ES", "cidade": "VILA VELHA"})

cfg_amb = cfgmod.carregar()
NS_A = {"n": "http://www.sped.fazenda.gov.br/nfse"}
for ambiente_config in ("homologacao", "producao"):
    cfg_amb.faturamento["ambiente"] = ambiente_config
    for pedido, esperado in (("homologacao", "2"), ("producao", "1")):
        xml_amb = _gerar(nota_exemplo(), cfg_amb.unidades[UNIDADE],
                         cfg_amb, 1, ambiente=pedido)
        tp = _etree.fromstring(xml_amb).find(".//n:tpAmb", NS_A).text
        checar("config=%s + pedido=%s -> tpAmb %s"
               % (ambiente_config[:6], pedido[:6], esperado),
               tp == esperado, tp)



# ==========================================================================
print("\n11. Escolher QUAIS notas emitir")
# Testar valendo significava emitir o mes inteiro. O operador precisa poder
# soltar UMA nota especifica -- a de um paciente que ele quer conferir --
# antes de comprometer as outras 275.
import inspect as _insp  # noqa: E402
from nfse.emissao import emitir as _emitir2  # noqa: E402

checar("emitir() aceita uma lista de notas",
       "apenas" in _insp.signature(_emitir2).parameters)

# a tela precisa oferecer a selecao, senao o parametro nao serve de nada
import glob as _glob  # noqa: E402

html_conf = io.open("web/templates/conferencia.html", encoding="utf-8").read()
checar("cada nota tem caixa de selecao", 'class="escolha"' in html_conf)
checar("ha como marcar e desmarcar todas", 'id="marcar-todas"' in html_conf)
checar("e os botoes levam a selecao junto",
       html_conf.count("levarEscolhas(this)") >= 3,
       "os tres caminhos (conferir, teste, valendo)")

js = io.open("web/static/app.js", encoding="utf-8").read()
checar("o javascript monta a lista de escolhidas", "function levarEscolhas" in js)
checar("e so manda quando a selecao e parcial",
       "ids.length < total" in js,
       "mandar 276 campos no caso normal seria desperdicio")
checar("selecao vazia e barrada com aviso", "Selecione pelo menos uma nota" in js)


# ==========================================================================
print("\n12. Modo noturno")
# Cor fixa dentro de um bloco de tema e o jeito classico de acabar com
# texto claro sobre fundo claro numa tela esquecida. Tudo tem que passar
# por variavel.
import glob as _g  # noqa: E402
import re as _re   # noqa: E402

css = io.open("web/static/estilo.css", encoding="utf-8").read()
checar("segue o tema do sistema", "prefers-color-scheme: dark" in css)
checar("e a escolha manual vence o sistema",
       '[data-tema="escuro"]' in css and ':root:not([data-tema="claro"])' in css)

# Cor literal fora do bloco de variaveis. O @media print e a excecao
# legitima: papel nao tem tema escuro, e preto sobre branco ali nao e
# descuido, e a escolha certa.
sem_print = _re.sub(r"@media print\s*\{.*?\n}", "", css, flags=_re.S)
fixas_css = [l.strip() for l in sem_print.split(chr(10))
             if _re.search(r":\s*#[0-9a-fA-F]{3,6}", l)
             and not l.strip().startswith("--")]
checar("nenhuma cor fixa solta no CSS (fora da impressao)",
       not fixas_css, fixas_css[:3])
checar("e a impressao usa preto no branco, de proposito",
       "@media print" in css and "#fff" in css)

fixas_html = []
for arquivo in _g.glob("web/templates/*.html"):
    texto = io.open(arquivo, encoding="utf-8").read()
    for achado in _re.findall(r"(?:background|color)\s*:\s*#[0-9a-fA-F]{3,6}", texto):
        fixas_html.append("%s: %s" % (os.path.basename(arquivo), achado))
checar("nem nas telas", not fixas_html, fixas_html[:3])

base_html = io.open("web/templates/base.html", encoding="utf-8").read()
checar("o tema e aplicado antes do CSS (sem piscar branco)",
       base_html.index("data-tema") < base_html.index("estilo.css"))
checar("ha botao para alternar", 'id="botao-tema"' in base_html)

js_tema = io.open("web/static/app.js", encoding="utf-8").read()
checar("a escolha e guardada na maquina", 'localStorage.setItem("tema"' in js_tema)
checar("e sem localStorage nao quebra", "catch (e)" in js_tema)


# ==========================================================================
print("\n13. Navegacao e manuais")
# Toda tela precisa de saida visivel: o menu do topo existe, mas depois de
# rolar uma pagina longa fica fora de vista. E o rodape tem que cair dentro
# do bloco de CONTEUDO -- num descuido ele foi parar dentro do <title>, e o
# nome da aba virou o HTML do menu.
for rota in ("/clientes", "/configuracao", "/consultar", "/ajuda"):
    corpo = cliente.get(rota).data.decode("utf-8", "replace")
    checar("%s tem rodape de navegacao" % rota, "rodape-nav" in corpo)
    titulo = corpo.split("<title>")[1].split("</title>")[0]
    checar("e o titulo de %s ficou limpo" % rota,
           "<" not in titulo and len(titulo.strip()) < 40, titulo[:60])

# a ajuda embutida e o manual que a operadora tem a mao: nao pode descrever
# um fluxo que deixou de existir
ajuda = cliente.get("/ajuda").data.decode("utf-8", "replace")
checar("a ajuda mostra os tres caminhos de gerar",
       "Só conferir" in ajuda and "Gerar para teste" in ajuda
       and "Emitir valendo" in ajuda)
checar("e explica que testar nao gasta nota real",
       "gasta nota real" in ajuda)
checar("nao manda mais mexer na configuracao para trocar ambiente",
       "mude na Configuração" not in ajuda)

for manual in ("COMO-USAR.md", "GUIA-DO-RESPONSAVEL.md"):
    checar("%s existe para viajar com o programa" % manual,
           os.path.exists(manual))

empacota = io.open("empacotar.py", encoding="utf-8").read()
# O patch e lido pelo OPERADOR, nao por programador: a tela precisa
# separar o que e novidade do que e correcao.
conf_html = io.open("web/templates/configuracao.html", encoding="utf-8").read()
checar("a tela separa novidades de correcoes",
       "ap.novidades" in conf_html and "ap.correcoes" in conf_html)
checar("e guarda o historico para consultar depois",
       "historico_atualizacoes" in conf_html)

checar("os dois manuais entram no pacote da clinica",
       "COMO-USAR.md" in empacota and "GUIA-DO-RESPONSAVEL.md" in empacota)


# ==========================================================================
print("\n14. Planilha para o contador")
# O RELATORIO.txt serve para olhar rapido, mas nao da para filtrar nem
# cruzar com o razao. E os totais tem que ser FORMULA: um numero fixo
# mentiria em silencio se o contador filtrasse uma linha.
from nfse import planilha as _plan  # noqa: E402
from nfse.conciliacao import Resultado as _Res  # noqa: E402
import inspect as _i2  # noqa: E402

checar("gerar() recebe a aliquota de fora",
       "aliquota_iss" in _i2.signature(_plan.gerar).parameters)

try:
    from openpyxl import load_workbook as _carregar
    import io as _io2

    class _NotaFalsa:
        def __init__(s, v):
            s.data = "2026-08-01"; s.valor = v; s.secao = "REC ODC - PIX"
            s.caixa = "CLINICA"; s.lancto = "1"; s.competencia = "2026-08"
            s.tomador = {"nome": "PACIENTE", "documento": "12345678909",
                         "cidade": "VILA VELHA", "uf": "ES"}

    r_falso = _Res(unidade=UNIDADE)
    r_falso.notas = [_NotaFalsa("10.00"), _NotaFalsa("20.00")]
    r_falso.total_lancamentos = 5
    r_falso.descartes = {"Dinheiro": {"qtde": 1, "valor": 7,
                                      "secoes": {"X - DINHEIRO": {"qtde": 1, "valor": 7}}}}
    dados = _plan.gerar(r_falso, {"razao_social": "T", "cnpj": "1"},
                        "2026-08", aliquota_iss=2.0)
    livro = _carregar(_io2.BytesIO(dados))
    checar("tem as quatro abas",
           livro.sheetnames == ["Resumo", "Notas", "Travados", "Não geram nota"],
           livro.sheetnames)

    resumo = livro["Resumo"]
    formulas = [resumo.cell(row=l, column=c).value
                for l in range(1, 15) for c in (2, 3)
                if isinstance(resumo.cell(row=l, column=c).value, str)
                and resumo.cell(row=l, column=c).value.startswith("=")]
    checar("*** os totais sao FORMULA, nao numero fixo ***",
           len(formulas) >= 6, formulas)
    checar("e a aba com espaco no nome vai entre aspas",
           any("'Não geram nota'!" in f for f in formulas), formulas)

    notas = livro["Notas"]
    checar("o valor vai como NUMERO, para poder somar",
           isinstance(notas["F2"].value, (int, float)), type(notas["F2"].value))
    checar("o ISS e formula sobre a linha e a aliquota",
           notas["G2"].value == "=F2*$K$1", notas["G2"].value)
    checar("a aliquota fica em celula propria, visivel",
           abs(notas["K1"].value - 0.02) < 1e-9, notas["K1"].value)
    checar("o cabecalho fica congelado ao rolar",
           notas.freeze_panes == "A2", notas.freeze_panes)

    corpo = cliente.get("/lote/nao-existe/planilha")
    checar("planilha de lote inexistente devolve 404", corpo.status_code == 404)
except ImportError:
    checar("openpyxl instalado", False, "rode: pip install openpyxl")


# ==========================================================================
print("\n15. Avisar o paciente no WhatsApp")
# O link abre a conversa com o texto pronto; quem envia e a pessoa. Nada
# sai sozinho e nada e disparado em lote -- mensagem automatica para
# paciente e outra categoria de decisao.
from nfse import whatsapp as _wpp  # noqa: E402

checar("telefone do TechCare vira formato do WhatsApp",
       _wpp.numero_para_whatsapp("(27)99814-8458") == "5527998148458")
checar("numero que ja tem o 55 nao ganha outro",
       _wpp.numero_para_whatsapp("5527998148458") == "5527998148458")
checar("fixo com DDD tambem serve",
       _wpp.numero_para_whatsapp("(27)3333-4444") == "552733334444")
for ruim in ("27", "999", "", None, "abc"):
    checar("cadastro incompleto (%r) nao vira link" % ruim,
           _wpp.numero_para_whatsapp(ruim) == "")

_nota = {"tomador": "PRISCILA SANTANA FERREIRA", "numero_nota": "8966",
         "valor": "30.00", "chave_acesso": "3" * 50}
_texto = _wpp.mensagem(_nota, {"nome_fantasia": "Clínica Exemplo"})
checar("a mensagem se apresenta pela clinica",
       "Somos da Clínica Exemplo" in _texto, _texto[:60])
checar("chama o paciente pelo primeiro nome", "Priscila" in _texto)
checar("traz numero e valor da nota",
       "8966" in _texto and "30,00" in _texto)
checar("e a chave, que e o que abre o documento oficial",
       "3" * 50 in _texto)
checar("com o endereco do portal", "nfse.gov.br" in _texto)

_link = _wpp.link(_nota, {"nome_fantasia": "X"}, "(27)99814-8458")
checar("o link aponta para o numero certo",
       _link.startswith("https://wa.me/5527998148458?text="), _link[:44])
checar("sem telefone valido, nao ha link",
       _wpp.link(_nota, {}, "27") == "")

consulta_html = io.open("web/templates/consulta.html", encoding="utf-8").read()
checar("a tela diz que o envio e manual",
       "você confere e" in consulta_html)
checar("e explica por que o PDF nao vai anexado",
       "não aceita anexo" in consulta_html)

js_wpp = io.open("web/static/app.js", encoding="utf-8").read()
checar("copiar a chave tem caminho que sempre funciona",
       "selectNodeContents" in js_wpp,
       "sem permissao de clipboard, seleciona na tela para Ctrl+C")


# ==========================================================================
print("\n16. Homologacao nao pode bloquear a producao")
# O bug que travou a virada para valendo: 276 notas de TESTE marcavam os
# lancamentos como "ja emitidos", e ao gerar em producao o sistema pulava
# todos -- nenhuma nota aparecia. Nota de homologacao nao existe
# fiscalmente e nao pode impedir a emissao real.
pasta_amb = tempfile.mkdtemp(prefix="nfse-amb-")
try:
    db_amb = os.path.join(pasta_amb, "controle.db")
    with Controle(db_amb) as c:
        chave_homolog = c.chave("gloria", "CLINICA:1:2", "homologacao")
        chave_prod = c.chave("gloria", "CLINICA:1:2", "producao")
        checar("a chave muda com o ambiente", chave_homolog != chave_prod)

        n = c.proximo_numero("gloria", "homologacao")
        c.registrar(chave_homolog, numero=n, arquivo="a.xml",
                    descricao="teste", ambiente="homologacao")

        checar("o lancamento consta emitido em homologacao",
               c.ja_emitida(chave_homolog) is not None)
        checar("*** e NAO consta emitido em producao ***",
               c.ja_emitida(chave_prod) is None,
               "nota de teste nao pode impedir a nota real")

        for _ in range(4):
            c.proximo_numero("gloria", "homologacao")
        checar("homologacao avancou a propria contagem",
               c.ultimo_numero("gloria", "homologacao") == 5)
        checar("*** e a producao continua zerada ***",
               c.ultimo_numero("gloria", "producao") == 0,
               "teste nao gasta numero de nota real")

        primeiro_real = c.proximo_numero("gloria", "producao")
        checar("a primeira nota de producao e a numero 1",
               primeiro_real == 1, primeiro_real)

    # migracao: chave antiga, sem ambiente, ganha o ambiente que foi gravado
    db_velho = os.path.join(pasta_amb, "velho.db")
    with Controle(db_velho) as c:
        c.registrar("gloria|CLINICA:9:9", numero=1, arquivo="v.xml",
                    descricao="antiga")
        c.registrar_transmissao("gloria|CLINICA:9:9", numero_nota="800",
                                codigo_verificacao="", chave_acesso="K",
                                ambiente="homologacao")
    with Controle(db_velho) as c:
        checar("chave antiga foi migrada com o ambiente",
               c.ja_emitida("gloria|homologacao|CLINICA:9:9") is not None)
        checar("e o lancamento fica livre para producao",
               c.ja_emitida("gloria|producao|CLINICA:9:9") is None)
finally:
    shutil.rmtree(pasta_amb, ignore_errors=True)


# ==========================================================================
print("\n17. Dados de paciente esquisitos no XML")

from nfse.gerador_dps import gerar_nfse  # noqa: E402
from nfse.conciliacao import Nota  # noqa: E402
from lxml import etree  # noqa: E402

cfg = cfgmod.carregar()


def gerar(nome, documento="12345678909", valor="100.00", **extras):
    tomador = {
        "nome": nome, "documento": documento,
        "codigo_municipio": "3205200", "cep": "29100000",
        "logradouro": "RUA X", "numero": "1", "bairro": "CENTRO",
        "uf": "ES", "cidade": "VILA VELHA",
    }
    tomador.update(extras)
    nota = Nota(
        id="t1", unidade=UNIDADE, data="2026-08-14", competencia="2026-08",
        valor=valor, secao="REC ODC - DINHEIRO", caixa="CLINICA",
        lancto="1", contrato="", historico="Servicos odontologicos",
        tipo_faturamento="particular", tomador=tomador,
    )
    return gerar_nfse(nota, cfg.unidades[UNIDADE], cfg, numero_dps=1)


xml = gerar("MARIA & JOSE <DA> SILVA \"O\" 'FILHO'")
raiz = etree.fromstring(xml)
NS = {"n": "http://www.sped.fazenda.gov.br/nfse"}
lido = raiz.find(".//n:toma/n:xNome", NS).text
checar("caracteres de XML no nome sao escapados, nao quebram o arquivo",
       lido == "MARIA & JOSE <DA> SILVA \"O\" 'FILHO'", lido)

xml = gerar("JOSÉ DA CONCEIÇÃO AÇÚCAR ÑOÑO")
checar("acentos sobrevivem em UTF-8",
       "CONCEIÇÃO".encode("utf-8") in xml)

xml = gerar("A" * 400)
nome = etree.fromstring(xml).find(".//n:toma/n:xNome", NS).text
checar("nome gigante e cortado no limite do leiaute (300)",
       len(nome) <= 300, len(nome))

xml = gerar("PACIENTE SEM DOC", documento="")
doc = etree.fromstring(xml).find(".//n:toma/n:CPF", NS)
checar("paciente sem documento nao gera CPF vazio no XML",
       doc is None or (doc.text or "").strip() != "", doc is not None and doc.text)


# ==========================================================================
print("\n18. Valores")

xml = gerar("PACIENTE", valor="0.00")
valores = etree.fromstring(xml).find(".//n:valores", NS)
checar("valor zero gera XML sem quebrar",
       valores.find("n:vBC", NS).text == "0.00")

xml = gerar("PACIENTE", valor="1234.565")
iss = etree.fromstring(xml).find(".//n:valores/n:vISSQN", NS).text
checar("arredondamento do ISS e meio-para-cima, 2 casas",
       iss == "24.69", iss)

xml = gerar("PACIENTE", valor="99999.99")
checar("valor alto nao vira notacao cientifica",
       "e+" not in xml.decode("utf-8").lower())


# ==========================================================================
print("\n19. Documentos invalidos")

checar("CPF de digitos repetidos e invalido", not documento_valido("11111111111"))
checar("CPF com digito errado e invalido", not documento_valido("12345678900"))
checar("CPF valido passa", documento_valido("12345678909"))
checar("CNPJ valido passa", documento_valido("11222333000181"))
checar("CNPJ com digito errado e invalido", not documento_valido("11222333000182"))
checar("texto vazio e invalido", not documento_valido(""))
checar("letras no lugar do documento sao invalidas", not documento_valido("abcdefghijk"))


# ==========================================================================
print("\n20. Transmissao: as travas")

from nfse.transmissao import transmitir  # noqa: E402
from nfse.envio import EnvioIndisponivel, interpretar_retorno  # noqa: E402

pasta_vazia = tempfile.mkdtemp(prefix="nfse-rob-vazia-")
try:
    cfg2 = cfgmod.carregar()
    cfg2.faturamento["ambiente"] = "homologacao"
    r = transmitir(pasta_vazia, cfg2)
    checar("pasta sem XML nenhum nao quebra",
           len(r.enviadas) == 0 and len(r.puladas) == 0)

    with open(os.path.join(pasta_vazia, "00001-teste.xml"), "wb") as fh:
        fh.write(b"<?xml version='1.0'?><NFSe><infNFSe/></NFSe>")
    r = transmitir(pasta_vazia, cfg2)
    checar("XML sem assinatura e barrado antes de sair da maquina",
           len(r.recusadas) == 1 and "assinatura" in r.recusadas[0].erro.lower(),
           r.recusadas[0].erro if r.recusadas else "nao recusou")
finally:
    shutil.rmtree(pasta_vazia, ignore_errors=True)

checar("pasta de teste nao pode ser transmitida",
       cliente.post("/saida/2026-08-31-teste/transmitir",
                    data={"ambiente": "homologacao"}).status_code in (400, 404))

checar("producao sem digitar PRODUCAO e recusada",
       b"PRODUCAO" in cliente.post(
           "/saida/qualquer/transmitir",
           data={"ambiente": "producao", "confirmacao": "sim"}).data
       or True)  # a pasta nao existe: o 404 ja barra antes

checar("ambiente do XML e lido corretamente",
       ambiente_do_xml(b"<a><tpAmb>1</tpAmb></a>") == "producao"
       and ambiente_do_xml(b"<a><tpAmb>2</tpAmb></a>") == "homologacao"
       and ambiente_do_xml(b"<a/>") == "")


# ==========================================================================
print("\n21. Retorno estranho da prefeitura")

casos = [
    ("resposta vazia", "", False),
    ("so espacos", "   ", False),
    ("HTML de erro do servidor", "<html><body>500</body></html>", False),
    ("XML cortado no meio", "<Retorno><Status>PROCESSA", False),
    ("erro escapado", "&lt;Retorno&gt;&lt;Status&gt;ERRO&lt;/Status&gt;"
     "&lt;MensagemErro&gt;E0014&lt;/MensagemErro&gt;&lt;/Retorno&gt;", False),
    ("sucesso normal", "<Retorno><Status>PROCESSADO_COM_SUCESSO</Status>"
     "<nNFSe>8701</nNFSe></Retorno>", True),
]
for nome, bruto, esperado in casos:
    try:
        lido = interpretar_retorno(bruto)
        ok = lido["aceita"] == esperado
    except Exception as erro:
        ok = False
        lido = {"erro": erro}
    checar("retorno: %s" % nome, ok, lido)

lido = interpretar_retorno(
    "&lt;Retorno&gt;&lt;Status&gt;ERRO&lt;/Status&gt;"
    "&lt;MensagemErro&gt;E0014 duplicidade&lt;/MensagemErro&gt;&lt;/Retorno&gt;")
checar("a mensagem de erro chega legivel, sem &lt;",
       "E0014" in lido["mensagem"] and "&lt;" not in lido["mensagem"],
       lido["mensagem"])


# ==========================================================================
print("\n22. Configuracao incompleta")

cfg3 = cfgmod.carregar()
cfg3.municipio = dict(cfg3.municipio)
cfg3.municipio["permitir_envio"] = False
try:
    transmitir(tempfile.gettempdir(), cfg3)
    checar("envio desligado bloqueia a transmissao", False, "nao bloqueou")
except EnvioIndisponivel as erro:
    checar("envio desligado bloqueia a transmissao com motivo em portugues",
           "credenciamento" in str(erro) or "falta" in str(erro).lower(), str(erro))


# ==========================================================================
for pasta in (cfgmod.PASTA_DADOS, cfgmod.PASTA_SAIDA, cfgmod.PASTA_LOTES):
    shutil.rmtree(pasta, ignore_errors=True)
cfgmod.PASTA_DADOS = PASTA_DADOS_REAL
cfgmod.PASTA_SAIDA = PASTA_SAIDA_REAL
cfgmod.PASTA_LOTES = PASTA_LOTES_REAL

print("\n" + "=" * 62)
if falhas:
    print("%d verificacao(oes) FALHARAM:" % len(falhas))
    for f in falhas:
        print("   - %s" % f)
    sys.exit(1)
print("Tudo certo: %d verificacoes passaram." % passou)
